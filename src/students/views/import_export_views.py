# students/views/import_export_views.py
import csv
import datetime
import io

import logging
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import ValidationError
from django.db import transaction
from django.core.cache import cache
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.views.generic import FormView, TemplateView, View

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..forms import *
from ..models import Parent, Student
from ..services.parent_service import ParentService
from ..services.student_service import StudentService
from ..utils import StudentUtils

logger = logging.getLogger(__name__)


class StudentImportView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """
    View for importing students from CSV/Excel files
    Updated to work without user account creation
    """

    form_class = BulkStudentImportForm
    permission_required = "students.bulk_import_students"
    template_name = "students/student_import.html"
    success_url = reverse_lazy("students:student-list")

    def form_valid(self, form):
        try:
            import_file = form.cleaned_data["import_file"]
            default_class = form.cleaned_data.get("default_class")
            send_welcome_emails = form.cleaned_data.get("send_welcome_emails", False)
            update_existing = form.cleaned_data.get("update_existing", False)

            # Determine file type
            if import_file.name.endswith(".csv"):
                result = self.process_csv_import(
                    import_file, default_class, send_welcome_emails, update_existing
                )
            elif import_file.name.endswith((".xlsx", ".xls")):
                result = self.process_excel_import(
                    import_file, default_class, send_welcome_emails, update_existing
                )
            else:
                messages.error(
                    self.request,
                    "Unsupported file format. Please use CSV or Excel files.",
                )
                return self.form_invalid(form)

            # Display results
            if result["created"] > 0:
                messages.success(
                    self.request, f"Successfully imported {result['created']} students."
                )

            if result["updated"] > 0:
                messages.success(
                    self.request, f"Successfully updated {result['updated']} students."
                )

            if result["errors"]:
                for error in result["errors"][:10]:  # Show first 10 errors
                    messages.error(
                        self.request, f"Row {error['row']}: {error['message']}"
                    )

                if len(result["errors"]) > 10:
                    messages.warning(
                        self.request,
                        f"And {len(result['errors']) - 10} more errors. Check the detailed report.",
                    )

            return super().form_valid(form)

        except Exception as e:
            logger.error(f"Error during student import: {str(e)}")
            messages.error(self.request, f"Import failed: {str(e)}")
            return self.form_invalid(form)

    def process_csv_import(
        self, csv_file, default_class, send_welcome_emails, update_existing
    ):
        """Process CSV file import"""
        result = {"created": 0, "updated": 0, "errors": []}

        try:
            csv_content = csv_file.read().decode("utf-8-sig")  # Handle BOM
            csv_reader = csv.DictReader(io.StringIO(csv_content))

            with transaction.atomic():
                for row_num, row in enumerate(csv_reader, start=2):
                    try:
                        student_result = self.process_student_row(
                            row,
                            row_num,
                            default_class,
                            send_welcome_emails,
                            update_existing,
                        )

                        if student_result["action"] == "created":
                            result["created"] += 1
                        elif student_result["action"] == "updated":
                            result["updated"] += 1

                    except Exception as e:
                        result["errors"].append({"row": row_num, "message": str(e)})
                        logger.error(f"Error processing row {row_num}: {str(e)}")

        except Exception as e:
            raise Exception(f"Error reading CSV file: {str(e)}")

        return result

    def process_excel_import(
        self, excel_file, default_class, send_welcome_emails, update_existing
    ):
        """Process Excel file import"""
        result = {"created": 0, "updated": 0, "errors": []}

        try:
            import openpyxl

            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(
                    cell.value.lower().replace(" ", "_") if cell.value else ""
                )

            with transaction.atomic():
                for row_num, row in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True), start=2
                ):
                    try:
                        # Convert row to dictionary
                        row_dict = {}
                        for i, value in enumerate(row):
                            if i < len(headers) and headers[i]:
                                row_dict[headers[i]] = value

                        student_result = self.process_student_row(
                            row_dict,
                            row_num,
                            default_class,
                            send_welcome_emails,
                            update_existing,
                        )

                        if student_result["action"] == "created":
                            result["created"] += 1
                        elif student_result["action"] == "updated":
                            result["updated"] += 1

                    except Exception as e:
                        result["errors"].append({"row": row_num, "message": str(e)})
                        logger.error(f"Error processing row {row_num}: {str(e)}")

        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")

        return result

    def process_student_row(
        self, row, row_num, default_class, send_welcome_emails, update_existing
    ):
        """
        Process a single student row from import file
        Updated to work without user account creation
        """
        # Clean and validate row data
        student_data = self.clean_row_data(row)

        # Validate required fields
        required_fields = [
            "first_name",
            "last_name",
            "admission_number",
            "emergency_contact_name",
            "emergency_contact_number",
        ]
        for field in required_fields:
            if not student_data.get(field):
                raise ValidationError(f"Missing required field: {field}")

        # Handle class assignment
        if not student_data.get("current_class") and default_class:
            student_data["current_class"] = default_class
        elif student_data.get("class_name"):
            try:
                student_data["current_class"] = Class.objects.get(
                    name=student_data["class_name"]
                )
            except Class.DoesNotExist:
                if default_class:
                    student_data["current_class"] = default_class
                else:
                    raise ValidationError(
                        f"Class '{student_data['class_name']}' not found"
                    )

        # Check if student already exists
        admission_number = student_data["admission_number"]
        existing_student = Student.objects.filter(
            admission_number=admission_number
        ).first()

        if existing_student:
            if update_existing:
                # Update existing student
                for field, value in student_data.items():
                    if (
                        field != "admission_number" and value
                    ):  # Don't update admission number
                        setattr(existing_student, field, value)

                existing_student.save()

                if send_welcome_emails and existing_student.email:
                    try:
                        StudentService.send_welcome_email(existing_student)
                    except Exception as e:
                        logger.error(f"Failed to send welcome email: {str(e)}")

                return {"action": "updated", "student": existing_student}
            else:
                raise ValidationError(
                    f"Student with admission number {admission_number} already exists"
                )
        else:
            # Create new student (without user account)
            student = StudentService.create_student(
                student_data, created_by=self.request.user
            )

            if send_welcome_emails and student.email:
                try:
                    StudentService.send_welcome_email(student)
                except Exception as e:
                    logger.error(f"Failed to send welcome email: {str(e)}")

            return {"action": "created", "student": student}

    def clean_row_data(self, row):
        """Clean and normalize row data for student creation"""
        cleaned_data = {}

        # Field mapping (CSV/Excel header -> model field)
        field_mapping = {
            "first_name": "first_name",
            "last_name": "last_name",
            "email": "email",
            "phone_number": "phone_number",
            "phone": "phone_number",
            "date_of_birth": "date_of_birth",
            "birth_date": "date_of_birth",
            "gender": "gender",
            "address": "address",
            "admission_number": "admission_number",
            "admission_date": "admission_date",
            "class_name": "class_name",
            "class": "class_name",
            "roll_number": "roll_number",
            "blood_group": "blood_group",
            "medical_conditions": "medical_conditions",
            "emergency_contact_name": "emergency_contact_name",
            "emergency_name": "emergency_contact_name",
            "emergency_contact_number": "emergency_contact_number",
            "emergency_phone": "emergency_contact_number",
            "emergency_contact_relationship": "emergency_contact_relationship",
            "previous_school": "previous_school",
            "status": "status",
        }

        for csv_field, model_field in field_mapping.items():
            value = row.get(csv_field)
            if value is not None:
                cleaned_data[model_field] = self.clean_field_value(model_field, value)

        return cleaned_data

    def clean_field_value(self, field_name, value):
        """Clean individual field values"""
        if value is None or str(value).strip() == "":
            return None

        value = str(value).strip()

        # Field-specific cleaning
        if field_name in ["first_name", "last_name", "emergency_contact_name"]:
            return value.title()

        elif field_name == "email":
            return value.lower() if value else None

        elif field_name == "admission_number":
            return value.upper()

        elif field_name == "gender":
            gender_map = {"male": "M", "female": "F", "m": "M", "f": "F"}
            return gender_map.get(value.lower(), value.upper()[:1])

        elif field_name == "blood_group":
            return StudentUtils.normalize_blood_group(value)

        elif field_name in ["date_of_birth", "admission_date"]:
            return self.parse_date(value)

        elif field_name == "status":
            status_map = {
                "active": "Active",
                "inactive": "Inactive",
                "graduated": "Graduated",
            }
            return status_map.get(value.lower(), value.title())

        return value

    def parse_date(self, date_value):
        """Parse date from various formats"""
        if not date_value:
            return None

        # Handle Excel date serial numbers
        if isinstance(date_value, (int, float)):
            try:
                # Excel date serial number
                from datetime import date, timedelta

                excel_epoch = date(1900, 1, 1)
                return excel_epoch + timedelta(days=int(date_value) - 2)
            except:
                return None

        # Handle string dates
        date_formats = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%m-%d-%Y",
            "%d-%m-%Y",
        ]

        for date_format in date_formats:
            try:
                return datetime.strptime(str(date_value), date_format).date()
            except ValueError:
                continue

        return None


class StudentExportView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """
    View for exporting student data
    Updated to export direct student fields instead of user fields
    """

    form_class = StudentExportForm
    permission_required = "students.export_student_data"
    template_name = "students/student_export.html"

    def form_valid(self, form):
        export_format = form.cleaned_data["export_format"]
        include_fields = form.cleaned_data.get("include_fields", [])
        class_filter = form.cleaned_data.get("class_filter")
        status_filter = form.cleaned_data.get("status_filter")

        # Build queryset
        queryset = Student.objects.all().select_related("current_class")

        if class_filter:
            queryset = queryset.filter(current_class=class_filter)

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Generate export
        if export_format == "csv":
            return self.export_csv(queryset, include_fields)
        elif export_format == "excel":
            return self.export_excel(queryset, include_fields)
        elif export_format == "json":
            return self.export_json(queryset, include_fields)
        else:
            messages.error(self.request, "Invalid export format")
            return self.form_invalid(form)

    def export_csv(self, queryset, include_fields):
        """Export students to CSV format"""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="students_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        )

        writer = csv.writer(response)

        # Write header
        headers = self.get_export_headers(include_fields)
        writer.writerow(headers)

        # Write data
        for student in queryset:
            row = self.get_student_export_row(student, include_fields)
            writer.writerow(row)

        return response

    def export_excel(self, queryset, include_fields):
        """Export students to Excel format"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Students"

        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

        # Write headers
        headers = self.get_export_headers(include_fields)
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_num, student in enumerate(queryset, 2):
            row_data = self.get_student_export_row(student, include_fields)
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="students_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )

        workbook.save(response)
        return response

    def export_json(self, queryset, include_fields):
        """Export students to JSON format"""
        data = []

        for student in queryset:
            student_data = {}

            # Get field mapping
            field_mapping = self.get_field_mapping()

            for field in include_fields:
                if field in field_mapping:
                    value = field_mapping[field](student)
                    # Handle date serialization
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()
                    student_data[field] = value

            data.append(student_data)

        response = HttpResponse(
            json.dumps(data, indent=2, default=str), content_type="application/json"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="students_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json"'
        )

        return response

    def get_export_headers(self, include_fields):
        """Get column headers for export"""
        header_mapping = {
            "admission_number": "Admission Number",
            "first_name": "First Name",
            "last_name": "Last Name",
            "full_name": "Full Name",
            "email": "Email",
            "phone_number": "Phone Number",
            "date_of_birth": "Date of Birth",
            "age": "Age",
            "gender": "Gender",
            "address": "Address",
            "admission_date": "Admission Date",
            "current_class": "Current Class",
            "roll_number": "Roll Number",
            "blood_group": "Blood Group",
            "medical_conditions": "Medical Conditions",
            "emergency_contact_name": "Emergency Contact Name",
            "emergency_contact_number": "Emergency Contact Number",
            "emergency_contact_relationship": "Emergency Contact Relationship",
            "previous_school": "Previous School",
            "status": "Status",
            "is_active": "Is Active",
            "date_joined": "Date Joined",
            "parents": "Parents",
        }

        return [header_mapping.get(field, field.title()) for field in include_fields]

    def get_student_export_row(self, student, include_fields):
        """Get data row for a student"""
        field_mapping = self.get_field_mapping()

        row = []
        for field in include_fields:
            if field in field_mapping:
                value = field_mapping[field](student)
                # Format value for CSV/Excel
                if value is None:
                    row.append("")
                elif hasattr(value, "strftime"):
                    row.append(value.strftime("%Y-%m-%d"))
                else:
                    row.append(str(value))
            else:
                row.append("")

        return row

    def get_field_mapping(self):
        """Get field mapping functions for export"""
        return {
            "admission_number": lambda s: s.admission_number,
            "first_name": lambda s: s.first_name,
            "last_name": lambda s: s.last_name,
            "full_name": lambda s: s.full_name,
            "email": lambda s: s.email or "",
            "phone_number": lambda s: s.phone_number or "",
            "date_of_birth": lambda s: s.date_of_birth,
            "age": lambda s: s.age,
            "gender": lambda s: s.get_gender_display() if s.gender else "",
            "address": lambda s: s.address or "",
            "admission_date": lambda s: s.admission_date,
            "current_class": lambda s: str(s.current_class) if s.current_class else "",
            "roll_number": lambda s: s.roll_number or "",
            "blood_group": lambda s: s.blood_group,
            "medical_conditions": lambda s: s.medical_conditions or "",
            "emergency_contact_name": lambda s: s.emergency_contact_name,
            "emergency_contact_number": lambda s: s.emergency_contact_number,
            "emergency_contact_relationship": lambda s: s.emergency_contact_relationship
            or "",
            "previous_school": lambda s: s.previous_school or "",
            "status": lambda s: s.status,
            "is_active": lambda s: "Yes" if s.is_active else "No",
            "date_joined": lambda s: s.date_joined,
            "parents": lambda s: ", ".join([p.full_name for p in s.get_parents()]),
        }


@login_required
@permission_required("students.view_student")
def download_import_template(request):
    """Download CSV template for student import"""
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        'attachment; filename="student_import_template.csv"'
    )

    writer = csv.writer(response)

    # Write headers with all possible fields
    headers = [
        "first_name",  # Required
        "last_name",  # Required
        "admission_number",  # Required
        "email",  # Optional
        "phone_number",  # Optional
        "date_of_birth",  # Optional (YYYY-MM-DD)
        "gender",  # Optional (M/F/O)
        "address",  # Optional
        "admission_date",  # Optional (YYYY-MM-DD, defaults to today)
        "class_name",  # Optional (will use default class if not provided)
        "roll_number",  # Optional
        "blood_group",  # Optional (A+, A-, B+, B-, AB+, AB-, O+, O-, Unknown)
        "medical_conditions",  # Optional
        "emergency_contact_name",  # Required
        "emergency_contact_number",  # Required
        "emergency_contact_relationship",  # Optional
        "previous_school",  # Optional
        "status",  # Optional (Active, Inactive, etc.)
    ]

    writer.writerow(headers)

    # Write example row
    example_row = [
        "John",  # first_name
        "Doe",  # last_name
        "STU-2024-ABC123",  # admission_number
        "john.doe@student.school.edu",  # email
        "+1234567890",  # phone_number
        "2010-05-15",  # date_of_birth
        "M",  # gender
        "123 Main St, City, State",  # address
        "2024-01-15",  # admission_date
        "5A",  # class_name
        "001",  # roll_number
        "A+",  # blood_group
        "No known allergies",  # medical_conditions
        "Jane Doe",  # emergency_contact_name
        "+1234567891",  # emergency_contact_number
        "Mother",  # emergency_contact_relationship
        "Previous Elementary School",  # previous_school
        "Active",  # status
    ]

    writer.writerow(example_row)

    return response


@csrf_exempt
@require_http_methods(["POST"])
@login_required
@permission_required("students.bulk_import_students")
def validate_import_file(request):
    """AJAX endpoint to validate import file before processing"""
    try:
        if "import_file" not in request.FILES:
            return JsonResponse({"valid": False, "message": "No file uploaded"})

        import_file = request.FILES["import_file"]

        # Check file size (5MB limit)
        if import_file.size > 5 * 1024 * 1024:
            return JsonResponse(
                {"valid": False, "message": "File size must be less than 5MB"}
            )

        # Check file format
        if not import_file.name.endswith((".csv", ".xlsx", ".xls")):
            return JsonResponse(
                {
                    "valid": False,
                    "message": "Invalid file format. Use CSV or Excel files.",
                }
            )

        # Validate file structure
        validation_result = validate_file_structure(import_file)

        return JsonResponse(validation_result)

    except Exception as e:
        return JsonResponse({"valid": False, "message": f"Validation error: {str(e)}"})


def validate_file_structure(import_file):
    """Validate the structure of import file"""
    try:
        if import_file.name.endswith(".csv"):
            return validate_csv_structure(import_file)
        else:
            return validate_excel_structure(import_file)
    except Exception as e:
        return {"valid": False, "message": f"File validation error: {str(e)}"}


def validate_csv_structure(csv_file):
    """Validate CSV file structure"""
    try:
        csv_content = csv_file.read().decode("utf-8-sig")
        csv_file.seek(0)  # Reset file pointer

        csv_reader = csv.DictReader(io.StringIO(csv_content))
        headers = csv_reader.fieldnames

        # Check for required headers
        required_headers = [
            "first_name",
            "last_name",
            "admission_number",
            "emergency_contact_name",
            "emergency_contact_number",
        ]
        missing_headers = [h for h in required_headers if h not in headers]

        if missing_headers:
            return {
                "valid": False,
                "message": f"Missing required columns: {', '.join(missing_headers)}",
            }

        # Count rows
        row_count = sum(1 for row in csv_reader)

        return {
            "valid": True,
            "message": f"File is valid. Found {row_count} student records.",
            "row_count": row_count,
            "headers": headers,
        }

    except Exception as e:
        return {"valid": False, "message": f"CSV validation error: {str(e)}"}


def validate_excel_structure(excel_file):
    """Validate Excel file structure"""
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active

        # Get headers
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(cell.value.lower().replace(" ", "_"))

        # Check for required headers
        required_headers = [
            "first_name",
            "last_name",
            "admission_number",
            "emergency_contact_name",
            "emergency_contact_number",
        ]
        missing_headers = [h for h in required_headers if h not in headers]

        if missing_headers:
            return {
                "valid": False,
                "message": f"Missing required columns: {', '.join(missing_headers)}",
            }

        # Count rows (excluding header)
        row_count = worksheet.max_row - 1

        return {
            "valid": True,
            "message": f"File is valid. Found {row_count} student records.",
            "row_count": row_count,
            "headers": headers,
        }

    except Exception as e:
        return {"valid": False, "message": f"Excel validation error: {str(e)}"}


class ParentBulkImportView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    template_name = "students/parent_bulk_import.html"
    form_class = ParentBulkImportForm
    permission_required = "students.bulk_import_parents"
    success_url = "/students/parents/"

    def form_valid(self, form):
        csv_file = form.cleaned_data["csv_file"]
        send_notifications = form.cleaned_data.get("send_email_notifications", False)
        update_existing = form.cleaned_data.get("update_existing", False)

        try:
            # Process the CSV file
            import_result = ParentService.bulk_import_parents(
                csv_file=csv_file,
                send_notifications=send_notifications,
                update_existing=update_existing,
                created_by=self.request.user,
            )

            if import_result["success"]:
                messages.success(
                    self.request,
                    f"Import completed! Created: {import_result['created']}, "
                    f"Updated: {import_result['updated']}, "
                    f"Errors: {import_result['errors']}",
                )

                if import_result["errors"] > 0 and import_result.get("error_details"):
                    error_summary = []
                    for error in import_result["error_details"][:5]:
                        error_summary.append(f"Row {error['row']}: {error['error']}")

                    if len(import_result["error_details"]) > 5:
                        error_summary.append(
                            f"... and {len(import_result['error_details']) - 5} more errors"
                        )

                    messages.warning(
                        self.request, f"Import errors: {'; '.join(error_summary)}"
                    )
            else:
                messages.error(
                    self.request,
                    f"Import failed: {import_result.get('error', 'Unknown error')}",
                )

        except Exception as e:
            messages.error(self.request, f"Import failed with error: {str(e)}")

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Bulk Import Parents"
        context["description"] = "Upload a CSV file to import multiple parents at once"

        # Add CSV template information
        context["csv_template_headers"] = [
            "first_name",
            "last_name",
            "email",
            "phone_number",
            "relation_with_student",
            "occupation",
            "workplace",
            "work_phone",
            "student_admission_number",
            "is_primary_contact",
            "can_pickup",
            "financial_responsibility",
        ]

        # Sample data for template
        context["csv_sample_data"] = [
            {
                "first_name": "Jane",
                "last_name": "Doe",
                "email": "jane.doe@example.com",
                "phone_number": "+1234567890",
                "relation_with_student": "Mother",
                "occupation": "Teacher",
                "student_admission_number": "ADM-2024-001",
                "is_primary_contact": "true",
            },
            {
                "first_name": "Bob",
                "last_name": "Smith",
                "email": "bob.smith@example.com",
                "phone_number": "+1234567891",
                "relation_with_student": "Father",
                "occupation": "Engineer",
                "student_admission_number": "ADM-2024-002",
                "is_primary_contact": "true",
            },
        ]

        return context


""" class ExportStudentsView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "students.export_student_data"

    def get(self, request, *args, **kwargs):
        # Get filter parameters
        filters = {
            "status": request.GET.get("status", ""),
            "class_id": request.GET.get("class", ""),
            "blood_group": request.GET.get("blood_group", ""),
            "admission_year": request.GET.get("admission_year", ""),
        }

        # Build queryset based on filters
        queryset = StudentService.search_students(
            query=request.GET.get("search", ""),
            filters={k: v for k, v in filters.items() if v},
        )

        # Generate CSV content
        csv_content = StudentService.export_students_to_csv(queryset)

        # Create HTTP response with CSV content
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="students_export_{timestamp}.csv"'
        )
        response.write(csv_content)

        return response """


class ExportParentsView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "students.export_parent_data"

    def get(self, request, *args, **kwargs):
        # Get filter parameters
        filters = {
            "relation": request.GET.get("relation", ""),
            "emergency_contact": request.GET.get("emergency_contact", ""),
        }

        # Build queryset based on filters
        queryset = Parent.objects.all().with_related()

        # Apply filters
        if filters["relation"]:
            queryset = queryset.filter(relation_with_student=filters["relation"])

        if filters["emergency_contact"]:
            queryset = queryset.filter(
                emergency_contact=filters["emergency_contact"] == "true"
            )

        # Apply search
        search_query = request.GET.get("search", "")
        if search_query:
            queryset = queryset.search(search_query)

        # Generate CSV content
        csv_content = ParentService.export_parents_to_csv(queryset)

        # Create HTTP response with CSV content
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="parents_export_{timestamp}.csv"'
        )
        response.write(csv_content)

        return response


class DownloadCSVTemplateView(LoginRequiredMixin, View):
    """Download CSV templates for import"""

    def get(self, request, template_type):
        if template_type == "students":
            headers = [
                "first_name",
                "last_name",
                "admission_number",  # Primary identifier
                "email",  # Optional
                "admission_date",
                "roll_number",
                "current_class_id",
                "blood_group",
                "emergency_contact_name",
                "emergency_contact_number",
                "status",
                "nationality",
                "religion",
                "city",
                "state",
                "country",
                "phone_number",
                "date_of_birth",
                "medical_conditions",
            ]
            filename = "student_import_template.csv"

            # Add sample data with emphasis on admission_number
            sample_data = [
                {
                    "first_name": "John",
                    "last_name": "Doe",
                    "admission_number": "ADM-2024-001",  # Required
                    "email": "john.doe@example.com",  # Optional
                    "admission_date": "2024-01-15",
                    "emergency_contact_name": "Jane Doe",
                    "emergency_contact_number": "+1234567890",
                    "status": "Active",
                    "blood_group": "O+",
                    "nationality": "Indian",
                    "city": "Mumbai",
                    "state": "Maharashtra",
                    "country": "India",
                },
                {
                    "first_name": "Alice",
                    "last_name": "Smith",
                    "admission_number": "ADM-2024-002",  # Required
                    "email": "",  # Can be empty
                    "admission_date": "2024-01-16",
                    "emergency_contact_name": "Bob Smith",
                    "emergency_contact_number": "+1234567891",
                    "status": "Active",
                    "blood_group": "A+",
                    "nationality": "Indian",
                    "city": "Delhi",
                    "state": "Delhi",
                    "country": "India",
                },
            ]

        elif template_type == "parents":
            headers = [
                "first_name",
                "last_name",
                "email",
                "phone_number",
                "relation_with_student",
                "occupation",
                "workplace",
                "work_phone",
                "student_admission_number",
                "is_primary_contact",
                "can_pickup",
                "financial_responsibility",
            ]
            filename = "parent_import_template.csv"

            # Add sample data
            sample_data = [
                {
                    "first_name": "Jane",
                    "last_name": "Doe",
                    "email": "jane.doe@example.com",
                    "phone_number": "+1234567890",
                    "relation_with_student": "Mother",
                    "occupation": "Teacher",
                    "student_admission_number": "ADM-2024-001",
                    "is_primary_contact": "true",
                    "can_pickup": "true",
                    "financial_responsibility": "true",
                }
            ]

        else:
            return JsonResponse({"error": "Invalid template type"}, status=400)

        # Create CSV response
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        writer = csv.DictWriter(response, fieldnames=headers)
        writer.writeheader()

        # Write sample data
        for row in sample_data:
            writer.writerow(row)

        return response


class ImportStatusView(LoginRequiredMixin, TemplateView):
    """View to check import status using AJAX"""

    template_name = "students/import_status.html"

    def get(self, request, *args, **kwargs):
        import_id = request.GET.get("import_id")

        if not import_id:
            return JsonResponse({"error": "Import ID required"}, status=400)

        # Get import status from cache (this would be set during import process)
        cache_key = f"import_status_{import_id}"
        status = cache.get(cache_key)

        if status is None:
            return JsonResponse({"error": "Import not found"}, status=404)

        return JsonResponse(status)


class BulkExportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Enhanced bulk export with multiple formats"""

    template_name = "students/bulk_export.html"
    permission_required = "students.export_student_data"

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get(
            "export_type"
        )  # students, parents, relationships
        format_type = request.POST.get("format", "csv")  # csv, xlsx, pdf
        include_photos = request.POST.get("include_photos") == "on"

        try:
            if export_type == "students":
                queryset = Student.objects.all().with_related().with_parents()

                # Apply filters from form
                filters = {}
                for key in ["status", "class", "blood_group", "admission_year"]:
                    value = request.POST.get(key)
                    if value:
                        filters[key] = value

                if filters:
                    queryset = StudentService.search_students("", filters)

                if format_type == "csv":
                    response = self._export_students_csv(queryset)
                elif format_type == "xlsx":
                    response = self._export_students_xlsx(queryset)
                elif format_type == "pdf":
                    response = self._export_students_pdf(queryset, include_photos)
                else:
                    messages.error(request, "Invalid export format")
                    return redirect(request.path)

            elif export_type == "parents":
                queryset = Parent.objects.all().with_related().with_students()

                # Apply filters
                relation = request.POST.get("relation")
                if relation:
                    queryset = queryset.filter(relation_with_student=relation)

                if format_type == "csv":
                    response = self._export_parents_csv(queryset)
                else:
                    messages.error(request, "Format not supported for parents export")
                    return redirect(request.path)

            else:
                messages.error(request, "Invalid export type")
                return redirect(request.path)

            return response

        except Exception as e:
            messages.error(request, f"Export failed: {str(e)}")
            return redirect(request.path)

    def _export_students_csv(self, queryset):
        csv_content = StudentService.export_students_to_csv(queryset)
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="students_export_{timestamp}.csv"'
        )
        response.write(csv_content)
        return response

    def _export_parents_csv(self, queryset):
        csv_content = ParentService.export_parents_to_csv(queryset)
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="parents_export_{timestamp}.csv"'
        )
        response.write(csv_content)
        return response

    def _export_students_xlsx(self, queryset):
        # Excel export implementation would go here
        # For now, fallback to CSV
        return self._export_students_csv(queryset)

    def _export_students_pdf(self, queryset, include_photos=False):
        # PDF export implementation would go here
        # For now, fallback to CSV
        return self._export_students_csv(queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["export_types"] = [
            ("students", "Students"),
            ("parents", "Parents"),
            ("relationships", "Student-Parent Relationships"),
        ]
        context["format_types"] = [
            ("csv", "CSV"),
            ("xlsx", "Excel (XLSX)"),
            ("pdf", "PDF"),
        ]
        return context
